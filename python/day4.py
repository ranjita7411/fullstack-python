# fobj=open("dravid.txt","r")

# strike_rates=[]

# for line in fobj:
#     columns=line.split()

#     runs=int(columns[2])
#     balls=int(columns[3])

#     strike_rate=runs/balls*100

#     strike_rates.append(strike_rate)

# import statistics as st
# deviation=st.stdev(strike_rates)
# print(deviation)

# import openpyxl
# print(openpyxl.__version__)

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Name"
ws["B1"] = "Marks"

ws["A2"] = "Vennela"
ws["B2"] = 95

wb.save("student.xlsx")

print("Excel file created successfully!")